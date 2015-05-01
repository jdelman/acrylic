# coding: utf-8
# ------------------------------------------------------------------------------
# Name:        ExcelRW.py
# Author:      emlazzarin
# ------------------------------------------------------------------------------

from collections import OrderedDict
from openpyxl import load_workbook, Workbook


class UnicodeReader(object):

    def __init__(self, filename, sheet_name_or_num=0):
        """
        Iterate quickly row-by-row through an Excel file.

        If you want to access a specific row, simply read the whole file with:

        rows = [row for row in UnicodeReader('myfile.xls')]

        ... and then access the intended row in the usual fashion.
        """
        self.__wb = load_workbook(filename=filename, read_only=True)
        self._sheet = None
        self.change_sheet(sheet_name_or_num)

    def change_sheet(self, sheet_name_or_num):
        """
        Calling this method changes the sheet in anticipation for the
        next time you create an iterator.

        If you change the active sheet while iterating on a UnicodeReader
        instance, it will continue to iterate correctly until completion.
        The next time you iterate through reader, it will begin all over
        again at whichever sheet you most recently changed to.
        """
        if isinstance(sheet_name_or_num, int):
            self._sheet = self.__wb[self.__wb.sheetnames[sheet_name_or_num]]
        elif isinstance(sheet_name_or_num, basestring):
            self._sheet = self.__wb[sheet_name_or_num]
        else:
            reason = "Must enter either sheet name or sheet number."
            raise Exception(reason)

    @property
    def numsheets(self):
        return len(self.__wb.sheetnames)

    @property
    def sheetnames(self):
        return self.__wb.sheetnames

    def __iter__(self):
        for row in self._sheet.rows:
            yield [cell.value for cell in row]


class UnicodeWriter(object):

    def __init__(self, filename):
        if not (isinstance(filename, basestring)):
            reason = ("UnicodeWriter requires a filename string"
                      " and not an open file.")
            raise Exception(reason)

        self.__wb = Workbook(write_only=True)
        self.__active_sheet = None
        self.__active_sheet_name = None
        self.__sheets = {}

        self.filename = filename

    def set_active_sheet(self, name):
        # Preserve the current sheet.
        self.__sheets[self.__active_sheet_name] = self.__active_sheet

        if name not in self.__sheets:
            ws = self.__wb.create_sheet()
            ws.title = name
            self.__sheets[name] = ws

        self.__active_sheet_name = name
        self.__active_sheet = self.__sheets[name]

    def writerow(self, row):
        if not self.__active_sheet:
            self.set_active_sheet("default")

        self.__active_sheet.append(row)

    def writerows(self, rows):
        if not self.__active_sheet:
            self.set_active_sheet("default")
        for row in rows:
            self.__active_sheet.append(row)

    def save(self):
        self.__wb.save(self.filename)

    def __exit__(self, exception_type, exception_value, traceback):
        self.save()


class UnicodeDictReader(UnicodeReader):

    def __init__(self, filename, sheet_name_or_num=0):
        super(UnicodeDictReader, self).__init__(filename, sheet_name_or_num)
 
    def __set_headers(self):
        self.__row_num = 0
        headers = super(UnicodeDictReader, self).__iter__().next()
 
        self._index_to_header = OrderedDict()
        for header in headers:
            self._index_to_header[headers.index(header)] = header
 
        self._header_indexes = sorted(self._index_to_header.keys())
 
        if range(len(headers)) != self._header_indexes:
            raise Exception("There appear to be hidden or 'extra'"
                            "columns in your *.xls file. This is "
                            "what ExcelRW sees:\n%s" % unicode(headers))
 
    def change_sheet(self, sheet_name_or_num):
        super(UnicodeDictReader, self).change_sheet(sheet_name_or_num)
        self.__set_headers()

    def __iter__(self):
        iterable = super(UnicodeDictReader, self).__iter__()
        iterable.next()  # burn off the headers
        for row in iterable:
            yield OrderedDict([(self._index_to_header[i], cell)
                               for i, cell in zip(self._header_indexes, row)])


class UnicodeDictWriter(UnicodeWriter):
    """
    A Excel writer that writes dicts in the order specified by fieldnames.

    Don't forget to call .save()!
    """

    def __init__(self, filename):
        super(UnicodeDictWriter, self).__init__(filename)
        self._cache = {}
        self._fieldnames = None
        self._headers = False
        self._sheetname = None

    def set_active_sheet(self, name, fieldnames=None):
        if name not in self._cache:
            if fieldnames is not None:
                self._cache[name] = self._fieldnames = fieldnames
            else:
                self._fieldnames = None
        else:
            if self._cache[name] != fieldnames and fieldnames is not None:
                values = (name, self._cache[name], fieldnames)
                raise Exception("Already begun writing to %s "
                                "with fieldnames %s. Cannot resume writing "
                                "with different fieldnames %s." % values)
            # in the case that the sheet name is in the cache
            # AND
            # either the fieldnames are the same as they were,
            # or if fieldnames is None
            # we load the fieldnames from cache
            self._fieldnames = self._cache[name]
        self._headers = False
        self._sheetname = name
        super(UnicodeDictWriter, self).set_active_sheet(name)

    @property
    def sheetname(self):
        return self._sheetname

    def writeheaders(self):
        if self._fieldnames:
            super(UnicodeDictWriter, self).writerow(self._fieldnames)
        else:
            self._headers = True

    def writerow(self, row):
        if self._fieldnames is None:
            self._cache[self._sheetname] = self._fieldnames = row.keys()
            # raise Exception("Fieldnames (headers) not defined.")

        if self._headers:
            super(UnicodeDictWriter, self).writerow(self._fieldnames)
            self._headers = False

        row_builder = []
        for fieldname in self._fieldnames:
            try:
                cell = row[fieldname]
            except KeyError:
                row_builder.append('')
                # raise Exception("Fieldname '%s' does not exist." % fieldname))
            else:
                row_builder.append(cell)
        super(UnicodeDictWriter, self).writerow(row_builder)

    def writerows(self, rows):
        for row_num, row in enumerate(rows):
            try:
                self.writerow(row)
            except Exception as e:
                print "Crashed on row %d:" % row_num
                print row
                raise e
