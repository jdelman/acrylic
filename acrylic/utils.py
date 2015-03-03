# coding: utf-8

from openpyxl import Workbook


def excel(path, datatables, sheetnames=None):
    wb = Workbook(write_only=True)
    if sheetnames is None:
        sheetnames = ["datatable_%02d" % i for i in range(1, len(datatables))]
    else:
        if len(sheetnames) != len(datatables):
            raise Exception("`sheetnames` is not the same "
                            "length as `datatables`: %s vs %s" %
                            (len(sheetnames), len(datatables)))
    for datatable, sheetname in zip(datatables, sheetnames):
        ws = wb.create_sheet()
        ws.title = sheetname
        ws.append(datatable.fields)
        for row in datatable:
            ws.append(row)
    wb.save(path)
